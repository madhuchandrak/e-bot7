# This is a utility file with the module to count the number of rows, number of columns, read and write to and from the same excel sheet

import openpyxl

def getRowCount(file,sheetname): # file is the parameter assigned to the excel sheet name and sheetname is the parameter assigned to the name of the sheet in which the data is present
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)

def getColumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet,max_column)

def readData(file,sheetname,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=colnum).value

def writeData(file,sheetname,rownum,colnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(file)